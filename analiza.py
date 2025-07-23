import pandas as pd
from openpyxl import load_workbook
import json # Necesario para manejar las respuestas JSON de la API de Gemini
import google.generativeai as genai

print("Iniciando la ejecución del script...") # Mensaje para confirmar que el script se está ejecutando

# --- Configuración de la API de Gemini ---
# IMPORTANTE: Reemplaza "XXXXXXXX" con tu clave de API real.
# NO COMPARTAS TU CLAVE DE API.
genai.configure(api_key="AIzaSyB-b6cWOUMEz4BjOIZDHRj-GYyqli9j2zY")

def analyze_with_gemini_placeholder(subject, description, resolution):
    """
    Esta función interactúa con la IA de Gemini para analizar un incidente.

    Args:
        subject (str): El asunto del incidente.
        description (str): La descripción del incidente.
        resolution (str): La resolución del incidente.

    Returns:
        dict: Un diccionario con 'es_candidato' (booleano) y 'razon' (string).
    """
    prompt = f"""Analiza el siguiente incidente y determina si es un candidato para una mejora definitiva. Proporciona una razón concisa.
Asunto: {subject}
Descripción: {description}
Resolución: {resolution}

Formato de respuesta JSON:
{{
  "es_candidato": boolean,
  "razon": "string"
}}
"""
    try:
        # Aquí está la llamada real a la API de Gemini.
        # La indentación ha sido corregida para que estas líneas estén dentro del bloque 'try'.
        model = genai.GenerativeModel('gemini-2.0-flash') # Puedes usar 'gemini-pro' o 'gemini-2.0-flash'
        response = model.generate_content(
            prompt,
            generation_config={
                "response_mime_type": "application/json" # Pedimos una respuesta en formato JSON
            }
        )

        # Intentamos parsear la respuesta de Gemini como JSON
        gemini_result = json.loads(response.text)

        return {
            "es_candidato": gemini_result.get("es_candidato", False),
            "razon": gemini_result.get("razon", "No se pudo obtener la razón del análisis de Gemini.")
        }

    except Exception as e:
        # Manejo de errores si la llamada a la API de Gemini falla
        print(f"Error al llamar a Gemini para el incidente '{subject}': {e}")
        return {"es_candidato": False, "razon": f"Error en la llamada a Gemini: {e}"}


def categorize_resolution(resolution_text):
    """
    Categoriza la resolución de un incidente basándose en palabras clave.
    """
    resolution_text_lower = str(resolution_text).lower()

    if "setok" in resolution_text_lower or "rerun" in resolution_text_lower:
        return "Setok/Rerun de Job"
    elif "timeout" in resolution_text_lower:
        return "Timeout de Proceso"
    elif "datos" in resolution_text_lower or "base de datos" in resolution_text_lower or "db" in resolution_text_lower:
        return "Problema de Datos/DB"
    elif "configuracion" in resolution_text_lower or "configuración" in resolution_text_lower:
        return "Problema de Configuración"
    elif "aplicacion" in resolution_text_lower or "aplicación" in resolution_text_lower or "job" in resolution_text_lower:
        return "Error en Aplicación/Job"
    elif "manual" in resolution_text_lower or "intervencion" in resolution_text_lower:
        return "Intervención Manual"
    elif "ok" in resolution_text_lower and ("finaliza" in resolution_text_lower or "resuelto" in resolution_text_lower):
        return "Resuelto OK (Genérico)"
    else:
        return "Otro Tipo de Resolución"


def analyze_incidents(file_path):
    """
    Lee un archivo Excel, procesa los incidentes y (con la implementación de Gemini)
    analiza si son candidatos para mejoras definitivas, escribiendo los resultados
    en una nueva pestaña del mismo archivo Excel.

    Args:
        file_path (str): La ruta completa al archivo Excel de entrada.
    """
    try:
        # Cargar el archivo Excel en un DataFrame de pandas
        df = pd.read_excel(file_path)

        # Filtrar solo las filas que corresponden a 'Incidente'
        incidents_df = df[df['Tipo de solicitud'] == 'Incidente'].copy()

        if incidents_df.empty:
            print("No se encontraron incidentes de tipo 'Incidente' en el archivo para analizar.")
            return

        # Asegurar que las columnas relevantes son de tipo string para el procesamiento
        incidents_df['Asunto'] = incidents_df['Asunto'].astype(str)
        incidents_df['Descripción'] = incidents_df['Descripción'].astype(str)
        incidents_df['Resolución'] = incidents_df['Resolución'].astype(str)

        # Listas para almacenar los resultados del análisis de Gemini y la nueva columna de agrupación
        gemini_es_candidato = []
        gemini_razon = []
        resolution_groups = [] # Nueva lista para almacenar los grupos de resolución

        print("Iniciando análisis de incidentes con Gemini (esto puede tomar tiempo si hay muchos tickets)...")
        # Iterar sobre cada incidente y pasarlo a la función de análisis de Gemini
        for index, row in incidents_df.iterrows():
            # Análisis con Gemini
            analysis_result = analyze_with_gemini_placeholder(
                row['Asunto'],
                row['Descripción'],
                row['Resolución']
            )
            gemini_es_candidato.append(analysis_result['es_candidato'])
            gemini_razon.append(analysis_result['razon'])

            # Categorización de la resolución para la nueva columna de agrupación
            resolution_groups.append(categorize_resolution(row['Resolución']))

            # Pequeño feedback para el usuario durante el procesamiento de grandes archivos
            if (index + 1) % 10 == 0:
                print(f"Procesados {index + 1} incidentes...")

        # Añadir las columnas de resultados de Gemini al DataFrame de incidentes
        incidents_df['Gemini_Es_Candidato_Mejora'] = gemini_es_candidato
        incidents_df['Gemini_Razon_Mejora'] = gemini_razon
        incidents_df['Grupo_Resolucion'] = resolution_groups # Añadir la nueva columna de agrupación

        # Seleccionar las columnas que se desean en la nueva pestaña.
        # Incluimos todas las columnas originales y las nuevas generadas por Gemini y la agrupación.
        # Aseguramos que el orden de las columnas sea el original más las nuevas al final.
        output_columns = df.columns.tolist() + ['Gemini_Es_Candidato_Mejora', 'Gemini_Razon_Mejora', 'Grupo_Resolucion']
        final_output_df = incidents_df[output_columns]

        # Definir el nombre de la nueva pestaña donde se guardarán los resultados
        output_sheet_name = 'Analisis_Gemini_Incidentes'

        # Cargar el libro de trabajo de Excel para poder manipular las hojas.
        # Esto es necesario para eliminar la pestaña si ya existe, evitando duplicados.
        book = load_workbook(file_path)
        if output_sheet_name in book.sheetnames:
            # Si la pestaña ya existe, se elimina para crearla de nuevo con los datos actualizados.
            del book[output_sheet_name]
        # Se guarda el libro de trabajo después de cualquier modificación (eliminación de hoja).
        book.save(file_path)

        # Usar pd.ExcelWriter en modo 'a' (append) para escribir en el archivo existente.
        # El motor 'openpyxl' es necesario para trabajar con archivos .xlsx.
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            # Escribir el DataFrame de resultados a la nueva pestaña, sin incluir el índice.
            final_output_df.to_excel(writer, sheet_name=output_sheet_name, index=False)

        print(f"Análisis completado. Los resultados se han guardado en la pestaña '{output_sheet_name}' del archivo '{file_path}'.")

    except FileNotFoundError:
        print(f"Error: El archivo '{file_path}' no fue encontrado. Por favor, asegúrate de que la ruta es correcta y el archivo existe.")
    except Exception as e:
        print(f"Ocurrió un error inesperado al procesar el archivo: {e}")

# --- Cómo usar esta aplicación ---
# 1. Guarda el código anterior en un archivo Python (por ejemplo, 'analizador_incidentes_gemini.py').
# 2. Asegúrate de tener las librerías 'pandas' y 'openpyxl' instaladas.
#    Si no las tienes, puedes instalarlas usando pip:
#    pip install pandas openpyxl
# 3. ¡IMPORTANTE! Instala la librería de Gemini:
#    pip install google-generativeai
# 4. Reemplaza 'tu_archivo_de_incidentes.xlsx' con la ruta real de tu archivo Excel.
#    Por ejemplo: file_path = 'C:/Usuarios/TuUsuario/Documentos/incidentes_resueltos.xlsx'
# 5. Asegúrate de que tu clave de API de Gemini esté configurada correctamente en la línea:
#    genai.configure(api_key="TU_CLAVE_API_DE_GEMINI")
# 6. Ejecuta el script desde tu terminal: python analizador_incidentes_gemini.py
#    El script modificará tu archivo Excel original, añadiendo una nueva pestaña con el análisis.

# --- Llamada principal para ejecutar el análisis ---
# Asegúrate de que la ruta del archivo sea correcta.
file_path = 'C:/Users/JoanManuelMunozRuiz/dev/vidasecurity/poc/incidentes/analisis_candidato_mantencion/data/2025_Tickets_muestra.xlsx'
analyze_incidents(file_path)
